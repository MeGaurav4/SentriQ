import uuid
import logging
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from sqlalchemy import select, text, MetaData
from app.db import engine, get_session
from app.models import ExcelImport

logger = logging.getLogger(__name__)

SERVER_FIELD_MAP = {
    "ip": "ip_address",
    "server_type": "server_type",
    "os": "os_version",
    "dc": "dc",
}


def _normalize(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def parse_wide_excel(filepath: str) -> tuple:
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows_iter = iter(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

    hostname_col = None
    collected_at_col = None
    param_cols = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        nh = _normalize(h)
        if nh == "hostname":
            hostname_col = i
        elif nh == "collected_at":
            collected_at_col = i
        else:
            param_cols[h] = i

    if hostname_col is None:
        raise ValueError("Missing required column: Hostname")

    data = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        hostname = str(row[hostname_col]).strip() if row[hostname_col] is not None else ""
        if not hostname:
            continue
        entry = {"__hostname__": hostname}
        if collected_at_col is not None:
            raw = row[collected_at_col]
            entry["__collected_at__"] = raw
        for col_name, col_idx in param_cols.items():
            entry[col_name] = row[col_idx]
        data.append(entry)

    return data, list(param_cols.keys()), headers


def _parse_collected_at(raw) -> datetime:
    now = datetime.now(timezone.utc)
    if raw is None:
        return now
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(raw))
        except (ValueError, TypeError, OverflowError):
            return now
    try:
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return now


def run_import(filepath: str, filename: str, os_family: str) -> dict:
    rows, param_columns, all_headers = parse_wide_excel(filepath)
    total = len(rows)
    success = 0
    errors = []
    warnings = []

    metadata = MetaData(schema="sq_schema")
    metadata.reflect(
        bind=engine, only=["servers", "parameter_definitions", "server_parameters"]
    )
    servers_t = metadata.tables["sq_schema.servers"]
    param_defs_t = metadata.tables["sq_schema.parameter_definitions"]
    server_params_t = metadata.tables["sq_schema.server_parameters"]

    os_family_lower = os_family.lower().strip()

    with engine.begin() as conn:
        all_param_defs = {}
        for row in conn.execute(
            select(param_defs_t.c.id, param_defs_t.c.name).where(
                param_defs_t.c.os_family == os_family_lower
            )
        ):
            all_param_defs[_normalize(row.name)] = row.id

        created_defs = {}
        all_servers = {}
        for row in conn.execute(
            select(servers_t.c.id, servers_t.c.hostname)
        ):
            all_servers[row.hostname.lower()] = row.id

        now = datetime.now(timezone.utc)
        insert_rows = []

        for i, entry in enumerate(rows):
            hostname = entry["__hostname__"]
            collected_at = _parse_collected_at(entry.get("__collected_at__"))

            server_id = all_servers.get(hostname.lower())
            if server_id is None:
                new_id = uuid.uuid4()
                extra_fields = {}
                for header in all_headers:
                    nh = _normalize(header)
                    if nh in SERVER_FIELD_MAP:
                        val = entry.get(header)
                        if val is not None:
                            extra_fields[SERVER_FIELD_MAP[nh]] = str(val).strip()
                ins = {
                    "id": new_id,
                    "hostname": hostname,
                    "os_family": os_family_lower,
                    "os_version": extra_fields.get("os_version", "Unknown"),
                    "server_type": (extra_fields.get("server_type") or "").lower().strip() or None,
                    "ip_address": extra_fields.get("ip_address"),
                    "is_baseline": False,
                    "last_seen": now,
                }
                conn.execute(servers_t.insert().values(**ins))
                all_servers[hostname.lower()] = new_id
                server_id = new_id

            for header in param_columns:
                raw_val = entry.get(header)
                if raw_val is None:
                    continue
                val = str(raw_val).strip()
                nh = _normalize(header)
                param_def_id = all_param_defs.get(nh)
                if param_def_id is None:
                    param_def_id = created_defs.get(nh)
                if param_def_id is None:
                    new_def_id = uuid.uuid4()
                    conn.execute(
                        text("""
                            INSERT INTO sq_schema.parameter_definitions
                                (id, name, display_name, os_family, data_type, category,
                                 is_active, sort_order, default_severity, help_text)
                            VALUES (:id, :name, :display_name, :os_family, :data_type,
                                    :category, :is_active, :sort_order, :default_severity, :help_text)
                            ON CONFLICT (name, os_family) DO NOTHING
                        """),
                        {
                            "id": new_def_id,
                            "name": nh,
                            "display_name": header,
                            "os_family": os_family_lower,
                            "data_type": "string",
                            "category": "system",
                            "is_active": True,
                            "sort_order": 0,
                            "default_severity": "warning",
                            "help_text": None,
                        },
                    )
                    all_param_defs[nh] = new_def_id
                    created_defs[nh] = new_def_id
                    param_def_id = new_def_id

                insert_rows.append({
                    "id": uuid.uuid4(),
                    "server_id": server_id,
                    "parameter_definition_id": param_def_id,
                    "parameter_value": val,
                    "collected_at": collected_at,
                    "source": "excel_import",
                })
                success += 1

        if insert_rows:
            conn.execute(server_params_t.insert(), insert_rows)

        import_id = uuid.uuid4()
        conn.execute(
            text("""
                INSERT INTO sq_schema.excel_imports
                    (id, filename, imported_at, row_count, success_count, error_count, error_log)
                VALUES (:id, :filename, :imported_at, :row_count, :success_count, :error_count, :error_log)
            """),
            {
                "id": import_id,
                "filename": filename,
                "imported_at": now,
                "row_count": total,
                "success_count": success,
                "error_count": len(errors),
                "error_log": "\n".join(errors[:100]) if errors else None,
            },
        )

    return {
        "import_id": import_id,
        "filename": filename,
        "total": total,
        "success": success,
        "errors": len(errors),
        "warnings": warnings[:100],
    }


def list_imports(limit: int = 20) -> list:
    with get_session() as session:
        return session.execute(
            select(ExcelImport)
            .order_by(ExcelImport.imported_at.desc())
            .limit(limit)
        ).scalars().all()
