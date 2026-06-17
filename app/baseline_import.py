import uuid
import logging
from datetime import datetime, timezone
from openpyxl import load_workbook
from sqlalchemy import select, text, MetaData
from app.db import engine

logger = logging.getLogger(__name__)

REQUIRED_COLS = {"server_type", "os_family"}


def _normalize(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def parse_wide_baseline_excel(filepath: str) -> tuple:
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows_iter = iter(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

    col_map = {}
    param_cols = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        nh = _normalize(h)
        if nh in ("server_type", "os_family", "os_version", "region"):
            col_map[nh] = i
        else:
            param_cols[h] = i

    missing = REQUIRED_COLS - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Missing required columns: {', '.join(sorted(missing))}. "
            f"Required: server_type, os_family"
        )

    data = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        entry = {}
        for col, idx in col_map.items():
            raw = row[idx] if idx < len(row) else None
            entry[col] = str(raw).strip() if raw is not None else ""
        for col_name, col_idx in param_cols.items():
            entry[col_name] = row[col_idx]
        entry["__param_cols__"] = list(param_cols.keys())
        if entry.get("server_type") and entry.get("os_family"):
            data.append(entry)

    return data, list(param_cols.keys()), headers


def run_baseline_import(filepath: str, filename: str, os_family_override: str = "") -> dict:
    rows, param_columns, all_headers = parse_wide_baseline_excel(filepath)
    total = len(rows)
    success = 0
    errors = []
    warnings = []

    metadata = MetaData(schema="sq_schema")
    metadata.reflect(
        bind=engine, only=["parameter_definitions", "baseline_configs"]
    )
    param_defs_t = metadata.tables["sq_schema.parameter_definitions"]
    baseline_t = metadata.tables["sq_schema.baseline_configs"]

    # Build param_def lookup per os_family
    with engine.begin() as conn:
        param_defs_by_os = {}
        for row in conn.execute(select(param_defs_t.c.id, param_defs_t.c.name, param_defs_t.c.os_family)):
            of = row.os_family.lower()
            if of not in param_defs_by_os:
                param_defs_by_os[of] = {}
            param_defs_by_os[of][_normalize(row.name)] = row.id

        created_defs = {}
        affected_combos = set()
        insert_rows = []
        now = datetime.now(timezone.utc)

        for i, entry in enumerate(rows):
            server_type = entry.get("server_type", "").lower().strip()
            os_family = entry.get("os_family", "").lower().strip()
            if os_family_override:
                os_family = os_family_override.lower().strip()
            os_version = entry.get("os_version", "") or None
            region = entry.get("region", "") or None

            if not server_type:
                errors.append(f"Row {i + 2}: empty server_type")
                continue
            if not os_family:
                errors.append(f"Row {i + 2}: empty os_family")
                continue

            if os_family not in param_defs_by_os:
                param_defs_by_os[os_family] = {}

            param_defs = param_defs_by_os[os_family]

            for header in entry.get("__param_cols__", param_columns):
                raw_val = entry.get(header)
                if raw_val is None:
                    continue
                val = str(raw_val).strip()
                nh = _normalize(header)
                param_def_id = param_defs.get(nh)
                if param_def_id is None:
                    param_def_id = created_defs.get(nh)
                if param_def_id is None:
                    new_def_id = uuid.uuid4()
                    conn.execute(
                        text("""
                            INSERT INTO sq_schema.parameter_definitions
                                (id, name, display_name, os_family, data_type, category,
                                 is_active, sort_order, default_severity, help_text)
                            VALUES (:id, :name, :display_name, :os_family, :data_type,
                                    :category, :is_active, :sort_order, :default_severity, :help_text)
                            ON CONFLICT (name, os_family) DO NOTHING
                        """),
                        {
                            "id": new_def_id,
                            "name": nh,
                            "display_name": header,
                            "os_family": os_family,
                            "data_type": "string",
                            "category": "system",
                            "is_active": True,
                            "sort_order": 0,
                            "default_severity": "warning",
                            "help_text": None,
                        },
                    )
                    param_defs[nh] = new_def_id
                    created_defs[nh] = new_def_id
                    param_def_id = new_def_id

                affected_combos.add((server_type, os_family, nh))

                is_critical = False
                insert_rows.append({
                    "id": uuid.uuid4(),
                    "server_type": server_type,
                    "os_family": os_family,
                    "os_version": os_version,
                    "region": region,
                    "category": "",
                    "parameter_key": header,
                    "parameter_definition_id": param_def_id,
                    "expected_value": val,
                    "is_critical": is_critical,
                    "updated_at": now,
                })
                success += 1

        if insert_rows:
            for st, of, pn in affected_combos:
                conn.execute(
                    text("""
                        DELETE FROM sq_schema.baseline_configs
                        WHERE LOWER(server_type) = :st
                          AND LOWER(os_family) = :of
                          AND LOWER(parameter_key) = :pn
                    """),
                    {"st": st, "of": of, "pn": pn},
                )
            conn.execute(baseline_t.insert(), insert_rows)

        import_id = uuid.uuid4()
        conn.execute(
            text("""
                INSERT INTO sq_schema.excel_imports
                    (id, filename, imported_at, row_count, success_count, error_count, error_log)
                VALUES (:id, :filename, :imported_at, :row_count, :success_count, :error_count, :error_log)
            """),
            {
                "id": import_id,
                "filename": filename,
                "imported_at": now,
                "row_count": total,
                "success_count": success,
                "error_count": len(errors),
                "error_log": "\n".join(errors[:100]) if errors else None,
            },
        )

    return {
        "import_id": import_id,
        "filename": filename,
        "total": total,
        "success": success,
        "errors": len(errors),
        "warnings": warnings[:100],
    }
