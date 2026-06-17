import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from app.db import get_session
from app.models import DriftReport, DriftFinding, Server


def generate_excel(report_id, output_dir):
    """
    Generates a color-coded Excel report for a specific DriftReport.
    """
    with get_session() as session:
        # 1. Fetch Report and Findings
        report = session.execute(
            select(DriftReport).where(DriftReport.id == report_id)
        ).scalar_one()

        findings = session.execute(
            select(DriftFinding, Server)
            .join(Server, DriftFinding.server_id == Server.id)
            .where(DriftFinding.report_id == report_id)
        ).all()

        # Prepare data for 'All Findings'
        all_findings_data = []
        for finding, server in findings:
            all_findings_data.append(
                {
                    "server": server.hostname,
                    "server_type": server.server_type,
                    "os": server.os_version,
                    "category": finding.category,
                    "key": finding.parameter_key,
                    "baseline": finding.baseline_value,
                    "current": finding.current_value,
                    "previous": finding.previous_value,
                    "type": finding.drift_type,
                    "severity": finding.severity,
                    "detected": finding.first_detected_at,
                }
            )

        wb = Workbook()

        # --- Sheet 1: Summary ---
        ws_sum = wb.active
        ws_sum.title = "Summary"
        headers = [
            "Report Date",
            "Role",
            "Servers",
            "Compliant",
            "Critical",
            "Warning",
            "Unreachable",
        ]
        ws_sum.append(headers)

        # Bold header
        for cell in ws_sum[1]:
            cell.font = Font(bold=True)
        ws_sum.freeze_panes = "A2"

        roles = sorted(list(set(f["server_type"] or "general" for f in all_findings_data)))

        for role in roles:
            role_servers = {f["server"] for f in all_findings_data if f["server_type"].lower() == role.lower()}
            role_crit = sum(
                1
                for f in all_findings_data
                if f["server_type"].lower() == role.lower() and f["severity"] == "critical"
            )
            role_warn = sum(
                1
                for f in all_findings_data
                if f["server_type"].lower() == role.lower() and f["severity"] == "warning"
            )
            # Simplified compliant calculation: servers in role minus those with critical/warning findings
            role_count = len(role_servers)

            row = [
                report.generated_at.strftime("%Y-%m-%d %H:%M"),
                role,
                role_count,
                "N/A",  # Compliant is harder to calculate per role without full server list
                role_crit,
                role_warn,
                "N/A",
            ]
            ws_sum.append(row)

            # Coloring
            last_row = ws_sum.max_row
            crit_cell = ws_sum.cell(row=last_row, column=5)
            warn_cell = ws_sum.cell(row=last_row, column=6)

            if role_crit > 0:
                crit_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            if role_warn > 0:
                warn_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )

        # --- Sheet 2: All Findings ---
        ws_all = wb.create_sheet("All Findings")
        headers_all = [
            "Server",
            "Role",
            "OS",
            "Category",
            "Parameter Key",
            "Baseline Value",
            "Current Value",
            "Previous Value",
            "Drift Type",
            "Severity",
            "First Detected",
        ]
        ws_all.append(headers_all)

        for cell in ws_all[1]:
            cell.font = Font(bold=True)
        ws_all.freeze_panes = "A2"

        # Colors
        fill_critical = PatternFill(
            start_color="DC2626", end_color="DC2626", fill_type="solid"
        )
        fill_warning = PatternFill(
            start_color="D97706", end_color="D97706", fill_type="solid"
        )
        font_white = Font(color="FFFFFF")

        for f in all_findings_data:
            row_vals = [
                f["server"],
                f["server_type"],
                f["os"],
                f["category"],
                f["key"],
                f["baseline"],
                f["current"],
                f["previous"],
                f["type"],
                f["severity"],
                f["detected"],
            ]
            ws_all.append(row_vals)

            row_idx = ws_all.max_row
            if f["severity"] == "critical":
                for cell in ws_all[row_idx]:
                    cell.fill = fill_critical
                    cell.font = font_white
            elif f["severity"] == "warning":
                for cell in ws_all[row_idx]:
                    cell.fill = fill_warning
                    cell.font = font_white

        # --- Role-based Sheets ---
        for role in roles:
            sheet_name = f"Role-{role}"[:31]
            ws_role = wb.create_sheet(sheet_name)
            ws_role.append(headers_all)
            for cell in ws_role[1]:
                cell.font = Font(bold=True)

            for f in all_findings_data:
                if f["server_type"].lower() == role.lower():
                    row_vals = [
                        f["server"],
                        f["server_type"],
                        f["os"],
                        f["category"],
                        f["key"],
                        f["baseline"],
                        f["current"],
                        f["previous"],
                        f["type"],
                        f["severity"],
                        f["detected"],
                    ]
                    ws_role.append(row_vals)
                    row_idx = ws_role.max_row
                    if f["severity"] == "critical":
                        for cell in ws_role[row_idx]:
                            cell.fill = fill_critical
                            cell.font = font_white
                    elif f["severity"] == "warning":
                        for cell in ws_role[row_idx]:
                            cell.fill = fill_warning
                            cell.font = font_white

        # --- OS-based Sheets ---
        os_versions = sorted(list(set(f["os"] for f in all_findings_data)))
        for os_v in os_versions:
            # Create a group name (e.g., Win-2019)
            group = "Unknown"
            if "2008" in os_v:
                group = "Win-2008"
            elif "2012" in os_v:
                group = "Win-2012"
            elif "2016" in os_v:
                group = "Win-2016"
            elif "2019" in os_v:
                group = "Win-2019"
            elif "2022" in os_v:
                group = "Win-2022"
            elif "RHEL" in os_v:
                group = "RHEL"

            sheet_name = group[:31]
            # Check if sheet already exists to avoid duplication
            if sheet_name in wb.sheetnames:
                continue

            ws_os = wb.create_sheet(sheet_name)
            ws_os.append(headers_all)
            for cell in ws_os[1]:
                cell.font = Font(bold=True)

            for f in all_findings_data:
                # This is a bit simplified; we check if it belongs to the group
                is_in_group = False
                if group == "Win-2008" and "2008" in f["os"]:
                    is_in_group = True
                elif group == "Win-2012" and "2012" in f["os"]:
                    is_in_group = True
                elif group == "Win-2016" and "2016" in f["os"]:
                    is_in_group = True
                elif group == "Win-2019" and "2019" in f["os"]:
                    is_in_group = True
                elif group == "Win-2022" and "2022" in f["os"]:
                    is_in_group = True
                elif group == "RHEL" and "RHEL" in f["os"]:
                    is_in_group = True

                if is_in_group:
                    row_vals = [
                        f["server"],
                        f["server_type"],
                        f["os"],
                        f["category"],
                        f["key"],
                        f["baseline"],
                        f["current"],
                        f["previous"],
                        f["type"],
                        f["severity"],
                        f["detected"],
                    ]
                    ws_os.append(row_vals)
                    row_idx = ws_os.max_row
                    if f["severity"] == "critical":
                        for cell in ws_os[row_idx]:
                            cell.fill = fill_critical
                            cell.font = font_white
                    elif f["severity"] == "warning":
                        for cell in ws_os[row_idx]:
                            cell.fill = fill_warning
                            cell.font = font_white

        # Auto-filter for findings sheets
        ws_all.auto_filter.ref = ws_all.dimensions

        # Filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"drift_report_{timestamp}.xlsx"
        full_path = os.path.join(output_dir, filename)

        os.makedirs(output_dir, exist_ok=True)
        wb.save(full_path)
        return full_path
